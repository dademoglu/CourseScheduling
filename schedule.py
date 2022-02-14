#By Deniz Ademoglu
#2019
# What it does:
#   There are two pages in data sheet where one has courses one has rooms, one workbook has only constraints. Before anything else it creates class element
#   for every item that is read from workbooks/sheets. Algorithm works like this:
#       take a course from the list
#       find it's constraints and check the other courses those are in that constraint group's schedule to find the hours that are already filled
#       take all the empty hours to place the course
#       compare every hour with rooms available times
#       place the course when first empty match is found
#       if couldn't find move on to the next room until you find a room with an appropiate empty slot
#
#This program doesn't guarantee eficiency in any way(memory, run time).
#This program doesn't guarantee that it will find empty places for all course
#This program doesn't guarantee to find the most suitiable room-course match


from openpyxl import load_workbook
from openpyxl import Workbook


class Room:
    M_full_hours = []
    T_full_hours = []
    def __init__(self, room_code, room_capacity, room_faculty):
        self.code = room_code
        self.capacity = room_capacity
        self.faculty = room_faculty
        self.M_free_hours = [1, 2, 3, 4, 5, 6]
        self.T_free_hours = [1, 2, 3, 4, 5, 6]
        self.F_free_hours = [1, 2, 3, 4, 5, 6]

    def __repr__(self):
        return "Faculty: %s | Room: %s | Capacity: %s |||####|||\n" % (self.faculty, self.code, self.capacity)

    def __str__(self):
        return "Faculty: %s | Room: %s | Capacity: %s" % (self.faculty, self.code, self.capacity)

class Course:
    placed = False
    def __init__(self, course_code, course_crn, course_section, course_title, course_faculty, course_capacity):
        self.code = course_code #SPS 303
        self.crn = course_crn #12312
        self.section = course_section #A
        self.title = course_title #Intro To Acc
        self.capacity = course_capacity #45
        self.faculty = course_faculty #FENS
        self.constraints = [] #[HIST 191, HIST 192]
        self.time = ''

    def __repr__(self):
        return "Course Code: %s | Course CRN: %s| Course Section: %s | Course Title: %s | Course Capacity: %s | Course Faculty: %s | Constraint Groups: %s|||####||| \n" % (self.code, self.crn, self.section, self.title, self.capacity, self.faculty, self.constraints)

    def __str__(self):
        return "Course Code: %s | Course Section: %s | Course Title: %s | Course Capacity: %s | Course Faculty: %s |||####|||" % (self.code, self.section, self.title, self.capacity, self.faculty)


class Schedule:
    def __init__(self, course_code, course_title, course_section, c_room, c_day, c_time, course_capacity2, room_capacity):
        self.c_code = course_code
        self.title = course_title
        self.section = course_section
        self.room = c_room
        self.day = c_day
        self.time = c_time
        self.r_cap = room_capacity
        self.c_cap = course_capacity2


    def __repr__(self):
        return "Course Title: %s | Course Section: %s | Course Room: %s | Course Day: %s | Course Time: %s | Capacity Efficiency: %s |||||####||| \n" % (self.title, self.section, self.room, self.day, self.time, (self.r_cap / self.c_cap))

    def __str__(self):
        return "Course Title: %s | Course Section: %s | Course Room: %s | Course Time: %s |||####||| \n" % (self.title, self.section, self.room, self.time)

class constraintGroup:
    def __init__(self, group_name):
        self.group = group_name
        self.group_elements = []
        self.scheduled = {
        "Monday 1":"",
        "Monday 2":"",
        "Monday 3":"",
        "Monday 4":"",
        "Monday 5":"",
        "Monday 6":"",
        "Tuesday 1":"",
        "Tuesday 2":"",
        "Tuesday 3":"",
        "Tuesday 4":"",
        "Tuesday 5":"",
        "Tuesday 6":""
        }
        self.scheduledF = {
        "Friday 1":"Empty",
        "Friday 2":"Empty",
        "Friday 3":"Empty",
        "Friday 4":"Empty",
        "Friday 5":"Empty",
        "Friday 6":"Empty"
        }


    def __repr__(self):
        return "Group Title: %s | Courses: %s |||||####||| \n" % (self.group, self.group_elements)

    def __str__(self):
        return "Group Title: %s | Courses: %s |||||####||| \n" % (self.group, self.group_elements)

def deleteTime(room, time, day):
        index = -1
        if(day == 'Monday'):
            if time in room.M_free_hours:
                index = room.M_free_hours.index(time)
                del room.M_free_hours[index]
        elif(day == 'Tuesday'):
            if time in room.T_free_hours:
                index = room.T_free_hours.index(time)
                del room.T_free_hours[index]



def findTime(course_name):
    empty_schedule = {
            "Monday 1":"Empty",
            "Monday 2":"Empty",
            "Monday 3":"Empty",
            "Monday 4":"Empty",
            "Monday 5":"Empty",
            "Monday 6":"Empty",
            "Tuesday 1":"Empty",
            "Tuesday 2":"Empty",
            "Tuesday 3":"Empty",
            "Tuesday 4":"Empty",
            "Tuesday 5":"Empty",
            "Tuesday 6":"Empty",
    }
    for constraint in constraint_groups:
        if course_name in constraint.group_elements:
            for time, course  in constraint.scheduled.items(): #fill an empty dictionary, to find the empty places for the course
                if course != "":
                    if(empty_schedule[time] != "Full"):
                        empty_schedule[time] = "Full"
                    else:
                        continue
        else:
            continue
    #print(empty_schedule)
    return empty_schedule

def findFridayTime(recit_name):
    empty_schedule = {
            "Friday 1":"Empty",
            "Friday 2":"Empty",
            "Friday 3":"Empty",
            "Friday 4":"Empty",
            "Friday 5":"Empty",
            "Friday 6":"Empty",
    }
    for constraint in constraint_recit_groups:
        if recit_name in constraint.group_elements:
            for time, course  in constraint.scheduledF.items(): #fill an empty dictionary, to find the empty places for the course
                if course != "":
                    if(empty_schedule[time] != "Full"):
                        empty_schedule[time] = "Full"
                    else:
                        continue
        else:
            continue
    #print(empty_schedule)
    return empty_schedule





wb = load_workbook('data.xlsx')
wb3 = load_workbook('Recitdata.xlsx')
recitbook = wb3['Recitations']
roombook = wb['Classes']
coursebook = wb['CourseNoRecRemade']
scheduleprogram = wb['Sheet1']

wb2 = load_workbook('constraints_2.xlsx')
wb4 = load_workbook('constraints_recits.xlsx')



rooms = []
recits = []
courses = []
constraints = {}
constraints_rec = {}
constraint_groups = []
constraint_recit_groups = []
firstroom = True
firstcourse = True
firstgroup = True
first_course_cell = True
second_course_cell = True
firstrecit = True
first_recit_cell = True
second_recit_cell = True
isaGroup = False
i = 0
crn = 0
group_title = "XX_XX_T"

for row in roombook.rows: #ROOMS ARE READ AT ONCE AND PLACED IN A LIST
    if(firstroom == True): #FIRST ROW IS HEADERS
        firstroom = False
        continue
    room_info = []
    for cell in row:
        room_info.append(cell.value)
    new_room = Room(room_info[1], room_info[2], room_info[0])
    rooms.append(new_room)

for row in coursebook.rows:             #COURSES ARE READ AND APPENDED TO LIST
    if(firstcourse == True):
        firstcourse = False
        continue
    course_info = []
    first_course_cell = True
    second_course_cell = True
    for cell in row:
        if (first_course_cell == True): #///////////////THIS PART IS TO MAKESURE CODE OF THE COURSE IS SHOWN TOGETHER
            temp_value = cell.value
            first_course_cell = False
            continue
        if (second_course_cell == True):#///////////////THIS PART IS TO MAKESURE CODE OF THE COURSE IS SHOWN TOGETHER
            temp_crn = cell.value
            course_info.append(str(temp_value) + " " + str(cell.value))
            second_course_cell = False
            continue
        course_info.append(cell.value)
    new_course = Course(course_info[0], course_info[1], course_info[2], course_info[3], course_info[4], course_info[5])
    if new_course.crn == crn:
        continue
    crn = new_course.crn
    courses.append(new_course)


for sheet in wb2: #CONSTRAINTS ARE DICTIONARIED
    firstgroup == True
    for row in sheet.rows:
        first_constraint_cell = True
        i = 0
        for cell in row:
            if(cell.value == "X"):
                if(firstgroup == False):
                    #print('yes')
                    constraints[group_title] = groups_list
                isaGroup = True
                #print('girdim')
                firstgroup = False
                groups_list = []
                continue
            elif(isaGroup == True):
                #print(cell.value)
                group_title = cell.value
                new_const_group = constraintGroup(group_title)
                constraint_groups.append(new_const_group)
                isaGroup = False

                break
            else:
                if(i == 2):
                    continue
                if(cell.value != None):
                    groups_list.append(cell.value)
                    new_const_group.group_elements.append(cell.value)
                    indices = [i for i, x in enumerate(courses) if x.code == cell.value]
                    for index in indices:
                        courses[index].constraints.append(cell.value)
            i += 1

#print(constraints)
#print(rooms)
#print('*****************************************************\n')
#print('/////////////////////////////////////////////////////\n')
#print(courses)

free_time = []
split_time = []
free_day = ''
free_hour = ''
schedule_list = []
schedule_list2 = []
available_time = []
count = 0
non_placed_courses = []
placed = False
no_match = False

for course in courses:
    placed = False
    available_time = findTime(course.code)
    for time, empty  in available_time.items():
        if empty == "Empty":
            free_time.append(time)
        else:
            non_placed_courses.append(course)
            continue
    if len(free_time) == 0:
        continue
    for room in rooms:
        ##I CAN IMPLEMENT room.capacity/course.capacity condition
        count = 0
        if placed == True:
            break
        for temp_time in free_time:
            split_time = temp_time.split()
            free_day = split_time[0]
            free_hour = int(split_time[1])
            if free_day == 'Monday' and free_hour in room.M_free_hours:
                room_to_add = room.faculty + " "+ room.code
                add_to_schedule = Schedule(course.code, course.title, course.section, room_to_add, "Monday", free_hour, course.capacity, room.capacity)
                placed = True
                del room.M_free_hours[room.M_free_hours.index(free_hour)]
                break
            elif free_day == 'Tuesday' and free_hour in room.T_free_hours:
                room_to_add = room.faculty + " "+ room.code
                add_to_schedule = Schedule(course.code, course.title, course.section, room_to_add, "Tuesday", free_hour, course.capacity, room.capacity)
                del room.T_free_hours[room.T_free_hours.index(free_hour)]
                placed = True
                break
            #count += 1
        #if count == len(free_time)
    schedule_list.append(add_to_schedule)


#/////////***********************************************RECITATIONS***********************************************/////////#
for row in recitbook.rows:             #RECITATIONS ARE READ AND APPENDED TO LIST
    if(firstrecit == True):
        firstrecit = False
        continue
    course_info = []
    first_recit_cell = True
    second_recit_cell = True
    for cell in row:
        if (first_recit_cell == True): #///////////////THIS PART IS TO MAKESURE CODE OF THE COURSE IS SHOWN TOGETHER
            temp_value = cell.value
            first_recit_cell = False
            continue
        if (second_recit_cell == True):#///////////////THIS PART IS TO MAKESURE CODE OF THE COURSE IS SHOWN TOGETHER
            temp_crn = cell.value
            course_info.append(str(temp_value) + " " + str(cell.value))
            second_recit_cell = False
            continue
        course_info.append(cell.value)
    new_recit = Course(course_info[0], course_info[1], course_info[2], course_info[3], course_info[4], course_info[5])
    if new_recit.crn == crn:
        continue
    crn = new_recit.crn
    recits.append(new_recit)



for sheet in wb4: #CONSTRAINTS FOR RECITATIONS ARE DICTIONARIED
    firstgroup == True
    for row in sheet.rows:
        first_constraint_cell = True
        i = 0
        for cell in row:
            if(cell.value == "X"):
                if(firstgroup == False):
                    #print('yes')
                    constraints_rec[group_title] = groups_list
                isaGroup = True
                #print('girdim')
                firstgroup = False
                groups_list = []
                continue
            elif(isaGroup == True):
                #print(cell.value)
                group_title = cell.value
                new_const_group = constraintGroup(group_title)
                constraint_recit_groups.append(new_const_group)
                isaGroup = False

                break
            else:
                if(i == 2):
                    continue
                if(cell.value != None):
                    groups_list.append(cell.value)
                    new_const_group.group_elements.append(cell.value)
                    indices = [i for i, x in enumerate(recits) if x.code == cell.value]
                    for index in indices:
                        exists = True
                        for cell.value in recits[index].constraints:
                            exists = False
                        if exists:
                            recits[index].constraints.append(cell.value)
            i += 1

free_time = []
split_time = []
free_day = ''
free_hour = ''
schedule_list = []
schedule_list2 = []
available_time = []
count = 0
placed = False
no_match = False
non_placed_recits = []
i = 0

for course in recits:
    #print(i)
    i += 1
    placed = False
    available_time = findFridayTime(course.code)
    for time, empty  in available_time.items():
        if empty == "Empty":
            free_time.append(time)
        else:
            non_placed_recits.append(course)
            continue
    if len(free_time) == 0:
        continue
    for room in rooms:
        ##I CAN IMPLEMENT room.capacity/course.capacity condition
        count = 0
        if placed == True:
            break
        for temp_time in free_time:
            split_time = temp_time.split()
            free_day = split_time[0]
            free_hour = int(split_time[1])
            if free_day == 'Friday' and free_hour in room.F_free_hours:
                room_to_add = room.faculty + " "+ room.code
                add_to_schedule = Schedule(course.code, course.title, course.section, room_to_add, "Friday", free_hour, course.capacity, room.capacity)
                placed = True
                del room.F_free_hours[room.F_free_hours.index(free_hour)]
                break
            #count += 1
        #if count == len(free_time)
    print(add_to_schedule.c_cap)
    schedule_list2.append(add_to_schedule)


#print(non_placed_courses)
print(len(non_placed_recits))
#print(schedule_list)
#print(schedule_dict)
#for item in constraint_groups:
#    print(str(item.group) + '\n')
#    print(item.scheduled)
#    print("////////////////////////////////////////////////////\n")

wb5 = Workbook()
ws3 = wb5.create_sheet("Schedule")
ws4 = wb5.create_sheet("Recitations")
ws3.insert_cols(6)
ws3.insert_rows(len(schedule_list))
ws4.insert_cols(6)
ws4.insert_rows(len(schedule_list))

for lecture in schedule_list:
        #print(lecture.room)
        ws3.append([lecture.c_code,lecture.title, lecture.section, lecture.room, lecture.day,lecture.time,(lecture.r_cap/lecture.c_cap)])

for lecture in schedule_list2:
        #print(lecture.room)
        ws4.append([lecture.c_code,lecture.title, lecture.section, lecture.room, lecture.day,lecture.time, lecture.r_cap, lecture.c_cap])


wb5.save("Schedule_Deniz_Ademoglu_2.xlsx")
